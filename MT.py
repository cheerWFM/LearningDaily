import random

from openpyxl import Workbook

time_title = ['*会议室名称','*类型','*可预约时段','容纳人数','*是否启用','备注']

b = Workbook()
c = b.active
c.title = "会议室导入模板"

c.merge_cells('A1:F1')
c["A1"]="注意事项"

txt = [time_title]
for row in txt:
	c.append(row)

for mod in range(1,500):
	if mod<10:
		value = "00" + str(mod)
		c.cell(row = mod+2,column = 1,value = value)
	else:
		value = "0" + str(mod)
		c.cell(row = mod+2,column = 1,value = value)

	
for mod2 in range(1,500):
	value = "会议室"
	c.cell(row = mod2+2,column = 2,value = value)
	
for mod3 in range(1,500):
	value = "00:00-24:00"
	c.cell(row = mod3+2,column = 3,value = value)
	
for mod4 in range(1,500):
	value = random.randint(1,50)
	c.cell(row = mod4+2,column = 4,value = value)

for mod5 in range(1,500):
	value = "是"
	c.cell(row = mod5+2,column = 5,value = value)
	
b.save("/Users/fangmeiwen/Test.xls")